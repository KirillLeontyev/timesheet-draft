from module_get_job import *
import openpyxl as xl


class CreateExcel:
    def __init__(self):
        self.services = ComposeJobList().compose()
        self.clients = {}
        self.total = {}
        logging.info(''.join((classname(self), 'Object initialized')))

    def write(self, filename, savename, services, client):
        # DEPRECATED
        # idc = client

        XL_CLIENTNAME = 'B1'
        XL_YEAR = 'G1'
        XL_MONTH = 'G2'
        XL_SERVICE_TOTAL = 'E'
        XL_TOTAL = 25
        XL_START_ROW = 6

        if client == 'npa':
            client = 'Неоплачиваемые'
            XL_TOTAL = 36
        else:
            client = client_fullname(client)[0]

        entries_to_write = []
        total_by_client = 0

        for index, service in enumerate(services):
            number = str(index + 1) + '.'
            date = service[0]
            start = service[1]
            end = service[2]
            total = end - start
            total_by_client += total
            desc = service[3]
            entries_to_write.append([number, date, time_format(start), time_format(end), time_format(total), desc])

        self.total.setdefault(id, total_by_client)

        files = get_xlsx()
        if savename in files:
            wb = xl.load_workbook(savename)
        else:
            wb = xl.load_workbook(filename)

        sheet_num = 0

        for index_service, service in enumerate(entries_to_write):
            # Calculating page_index to know when it's time to switch to the next sheet
            page_index = index_service
            coef = XL_TOTAL - XL_START_ROW

            # If N of entry > than rows in current sheet
            if page_index > (coef - 1):
                # Zeroing of page_index for next sheet
                page_index = index_service % coef
                if page_index == 0:
                    sheet_num += 1
                if sheet_num > 2:
                    sheet_num = 2
                    logging.error(''.join((classname(self), 'Sheet index exceeds total number of sheets')))

            # Placing a row with service
            for index_entry, entry in enumerate(service):
                ws = wb.worksheets[sheet_num]

                if page_index == 0:
                    ws[XL_CLIENTNAME] = client
                    ws[XL_YEAR] = date.year
                    ws[XL_MONTH] = MONTHS_NAMES[date.month - 1]
                    ws[XL_SERVICE_TOTAL + str(XL_TOTAL)] = time_format(total_by_client)

                ws.cell(page_index + XL_START_ROW, index_entry + 1, entry)
            print()

        wb.save(savename)

    def get_filename(self, services, client):
        date = services[0][0]

        filename = 'TIME 1_TEMPLATE_MM_YYYY.xlsx'
        savename = 'TIME 1_' + client + '_' + str(date.month) + '_' + str(date.year) + '.xlsx'
        if client == 'npa':
            filename = 'TIME 1_NPA_MM_YYYY.xlsx'
            savename = 'TIME 1_Неоплачиваемые_' + str(date.month) + '_' + str(date.year) + '.xlsx'

        try:
            self.write(filename, savename, services, client)
        except PermissionError:
            logging.error(''.join((classname(self), 'Excel file is already in use by another program')))

    def start(self):
        for day, hours in self.services:
            for service in day:
                client = client_fullname(str(service[4]))[1]
                service[4] = client
                if service[4]:
                    key = client
                else:
                    key = 'npa'
                self.clients.setdefault(key, [])
                self.clients[key].append(service)

        for client, services in self.clients.items():
            self.get_filename(services, client)
